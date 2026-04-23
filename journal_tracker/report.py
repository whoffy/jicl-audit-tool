"""Excel report with color-coded footnote ranges matching JICL Editing Audit format."""

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BOLD = Font(name="Arial", bold=True, size=11)
SECTION_FONT = Font(name="Arial", bold=True, size=13)
SECTION_FILL = PatternFill(start_color="C9DAF8", end_color="C9DAF8", fill_type="solid")
DIFFICULT_FILL = PatternFill(start_color="D9D2E9", end_color="D9D2E9", fill_type="solid")
THIN_BOTTOM = Border(bottom=Side(style="thin"))
THIN_TOP = Border(top=Side(style="thin"))
HEADER_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
RIGHT = Alignment(horizontal="right")

TAB_COLORS = {
    "Combined Rankings": "4472C4",
    "analytics": "A5A5A5",
}
TEAM_TAB_COLORS = [
    ("4472C4", "6C9BD2"),
    ("548235", "7BAF4E"),
    ("BF8F00", "D4A843"),
    ("C55A11", "D8844A"),
    ("7030A0", "9557B8"),
    ("2E75B6", "5B9BD5"),
]

PALETTE = [
    "F4CCCC",
    "FFF2CC",
    "D9EAD3",
    "C9DAF8",
    "D9D2E9",
    "CFE2F3",
    "EAD1DC",
]


def _fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _assign_colors(assignments):
    if not assignments:
        return [], {}
    people = sorted(assignments, key=lambda p: assignments[p]["fn_start"])
    colors = {p: _fill(PALETTE[i % len(PALETTE)]) for i, p in enumerate(people)}
    return people, colors


def _range_fill(fn, assignments, colors):
    for person, a in assignments.items():
        if a["fn_start"] <= fn <= a["fn_end"]:
            return colors.get(person)
    return None


def _finalize(ws):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                old = cell.font
                cell.font = Font(
                    name="Arial",
                    bold=old.bold,
                    size=old.size if old.size else 11,
                )
                if isinstance(cell.value, (int, float)) and not cell.alignment.horizontal:
                    cell.alignment = RIGHT
    _auto_width(ws)


def write_full_report(sp_metrics, sp_leaderboards, sp_detail,
                      ed_metrics, ed_details, analytics,
                      sp_assignments_by_team, ed_assignments_by_team,
                      output_path):
    wb = Workbook()
    _write_combined_rankings(wb, sp_leaderboards, ed_metrics)
    _write_team_sheets(wb, sp_metrics, sp_detail, ed_metrics, ed_details,
                       sp_assignments_by_team, ed_assignments_by_team, analytics)
    _write_analytics_sheets(wb, analytics)
    wb.save(output_path)


def _write_combined_rankings(wb, sp_leaderboards, ed_metrics):
    ws = wb.active
    ws.title = "Combined Rankings"
    ws.sheet_properties.tabColor = TAB_COLORS["Combined Rankings"]

    # --- Top section headers (row 3) ---
    for c in (2, 3, 4):
        ws.cell(row=3, column=c).fill = SECTION_FILL
    ws.cell(row=3, column=2, value="Source Pull Leaderboard").font = SECTION_FONT
    for c in (9, 10, 11):
        ws.cell(row=3, column=c).fill = SECTION_FILL
    ws.cell(row=3, column=9, value="Editing Leaderboard").font = SECTION_FONT
    ws.row_dimensions[3].height = 25.5

    # --- Sub-headers (row 5, gap after section header) ---
    for col, val in [(2, "Total Contributions"), (3, "Name"), (4, "#")]:
        ws.cell(row=5, column=col, value=val).border = THIN_BOTTOM
    for col, val in [(9, "Highest Total Contributions"), (10, "Name"), (11, "#")]:
        ws.cell(row=5, column=col, value=val).border = THIN_BOTTOM

    # SP Top 5 (rows 6-10)
    if sp_leaderboards:
        for i, (_, r) in enumerate(sp_leaderboards["highest_contributions"].head(5).iterrows()):
            ws.cell(row=6 + i, column=2, value=i + 1)
            ws.cell(row=6 + i, column=3, value=r["person"])
            ws.cell(row=6 + i, column=4, value=r["total"])

    # ED Top 6 (rows 6-11)
    if ed_metrics is not None:
        agg = ed_metrics.groupby("person", as_index=False)["total"].sum()
        top = agg.nlargest(6, "total").reset_index(drop=True)
        for i, (_, r) in enumerate(top.iterrows()):
            ws.cell(row=6 + i, column=9, value=i + 1)
            ws.cell(row=6 + i, column=10, value=r["person"])
            ws.cell(row=6 + i, column=11, value=r["total"])

    # --- SP quality / ED carries (row 13, gap after data) ---
    ws.cell(row=13, column=2,
            value="Mistakes in Assigned (Other Editor Edits to Assigned / Total Contributions)"
            ).border = THIN_BOTTOM
    if sp_leaderboards:
        bq = sp_leaderboards["best_quality"].head(9)
        for i, (_, r) in enumerate(bq.iterrows()):
            ws.cell(row=14 + i, column=2, value=i + 1)
            ws.cell(row=14 + i, column=3, value=r["person"])
            cell = ws.cell(row=14 + i, column=4, value=r["quality_ratio"])
            cell.number_format = '0.00%'

    ws.cell(row=13, column=9,
            value="Biggest Carries (Share of Team's Edits)"
            ).border = THIN_BOTTOM
    if ed_metrics is not None:
        carries = ed_metrics.nlargest(5, "share_of_team").reset_index(drop=True)
        for i, (_, r) in enumerate(carries.iterrows()):
            ws.cell(row=14 + i, column=9, value=i + 1)
            ws.cell(row=14 + i, column=10, value=r["person"])
            cell = ws.cell(row=14 + i, column=11, value=r["share_of_team"])
            cell.number_format = '0.00%'

    # --- Fewest section headers (row 25, gap after quality data) ---
    f_row = 25
    for c in (2, 3, 4):
        ws.cell(row=f_row, column=c).fill = SECTION_FILL
    ws.cell(row=f_row, column=2, value="Source Pull Fewest Contributions").font = SECTION_FONT
    for c in (9, 10, 11):
        ws.cell(row=f_row, column=c).fill = SECTION_FILL
    ws.cell(row=f_row, column=9, value="Editing Fewest Contributions").font = SECTION_FONT
    ws.row_dimensions[f_row].height = 25.5

    # Sub-headers (row 27, gap after section header)
    for col, val in [(2, "Total Contributions"), (3, "Name"), (4, "#")]:
        ws.cell(row=f_row + 2, column=col, value=val).border = THIN_BOTTOM
    for col, val in [(9, "Total Contributions"), (10, "Name"), (11, "#")]:
        ws.cell(row=f_row + 2, column=col, value=val).border = THIN_BOTTOM

    # Fewest 5 (rows 28-32)
    if sp_leaderboards:
        for i, (_, r) in enumerate(sp_leaderboards["fewest_contributions"].iterrows()):
            ws.cell(row=f_row + 3 + i, column=2, value=i + 1)
            ws.cell(row=f_row + 3 + i, column=3, value=r["person"])
            ws.cell(row=f_row + 3 + i, column=4, value=r["total"])

    if ed_metrics is not None:
        agg = ed_metrics.groupby("person", as_index=False)["total"].sum()
        bottom = agg.nsmallest(5, "total").reset_index(drop=True)
        for i, (_, r) in enumerate(bottom.iterrows()):
            ws.cell(row=f_row + 3 + i, column=9, value=i + 1)
            ws.cell(row=f_row + 3 + i, column=10, value=r["person"])
            ws.cell(row=f_row + 3 + i, column=11, value=r["total"])

    # --- Fewest sub-leaderboards (row 35, gap after data) ---
    sub_row = 35
    ws.cell(row=sub_row, column=2,
            value="Mistakes in Assigned (Other Editor Edits to Assigned / Total Contributions)"
            ).border = THIN_BOTTOM
    ws.cell(row=sub_row, column=9,
            value="Share of Team's Edits"
            ).border = THIN_BOTTOM

    if sp_leaderboards:
        wq = sp_leaderboards["worst_quality"]
        for i, (_, r) in enumerate(wq.iterrows()):
            ws.cell(row=sub_row + 1 + i, column=2, value=i + 1)
            ws.cell(row=sub_row + 1 + i, column=3, value=r["person"])
            cell = ws.cell(row=sub_row + 1 + i, column=4, value=r["quality_ratio"])
            cell.number_format = '0.00%'

    if ed_metrics is not None:
        lowest_share = ed_metrics.nsmallest(5, "share_of_team").reset_index(drop=True)
        for i, (_, r) in enumerate(lowest_share.iterrows()):
            ws.cell(row=sub_row + 1 + i, column=9, value=i + 1)
            ws.cell(row=sub_row + 1 + i, column=10, value=r["person"])
            cell = ws.cell(row=sub_row + 1 + i, column=11, value=r["share_of_team"])
            cell.number_format = '0.00%'

    # --- Most Difficult Articles (row 43, purple fill, right side) ---
    d_row = 43
    for c in (9, 10, 11):
        ws.cell(row=d_row, column=c).fill = DIFFICULT_FILL
    ws.cell(row=d_row, column=9, value="Most Difficult Articles").font = SECTION_FONT
    ws.row_dimensions[d_row].height = 25.5

    if ed_metrics is not None:
        for col, val in [(9, "Total Contributions"), (10, "Team Name"), (11, "#")]:
            ws.cell(row=d_row + 2, column=col, value=val).border = THIN_BOTTOM
        totals = ed_metrics.groupby("team", as_index=False)["total"].sum()
        totals = totals.sort_values("total", ascending=False).reset_index(drop=True)
        last_data_row = d_row + 3
        for i, (_, r) in enumerate(totals.iterrows()):
            last_data_row = d_row + 3 + i
            ws.cell(row=last_data_row, column=9, value=i + 1)
            ws.cell(row=last_data_row, column=10, value=r["team"])
            ws.cell(row=last_data_row, column=11, value=r["total"])
        for c in (9, 10, 11):
            ws.cell(row=last_data_row, column=c).border = THIN_BOTTOM

    _finalize(ws)


def _write_team_sheets(wb, sp_metrics, sp_detail, ed_metrics, ed_details,
                       sp_assignments_by_team, ed_assignments_by_team, analytics):
    teams = set()
    if sp_metrics is not None:
        teams.update(sp_metrics["team"].unique())
    if ed_metrics is not None:
        teams.update(ed_metrics["team"].unique())

    for idx, team in enumerate(sorted(teams)):
        sp_color, ed_color = TEAM_TAB_COLORS[idx % len(TEAM_TAB_COLORS)]

        if sp_metrics is not None and team in sp_metrics["team"].values:
            ws = wb.create_sheet(title=f"{team} Source Pull"[:31])
            ws.sheet_properties.tabColor = sp_color
            team_sp = sp_metrics[sp_metrics["team"] == team]
            det = (sp_detail[sp_detail["team"] == team]
                   .sort_values(["footnote", "modified_by"])
                   .reset_index(drop=True) if sp_detail is not None else pd.DataFrame())
            assigns = sp_assignments_by_team.get(team, {})
            _write_sp_team_sheet(ws, team_sp, det, assigns)

        if ed_metrics is not None and team in ed_metrics["team"].values:
            ws = wb.create_sheet(title=f"{team} Editing"[:31])
            ws.sheet_properties.tabColor = ed_color
            team_ed = ed_metrics[ed_metrics["team"] == team]
            det = ed_details.get(team, pd.DataFrame())
            assigns = ed_assignments_by_team.get(team, {})
            team_analytics = _filter_team_analytics(analytics, team)
            _write_ed_team_sheet(ws, team_ed, det, assigns, team_analytics)


def _filter_team_analytics(analytics, team):
    result = {}
    if not analytics:
        return result
    if "comment_quality" in analytics and not analytics["comment_quality"].empty:
        cq = analytics["comment_quality"]
        result["comment_quality"] = cq[cq["team"] == team]
    if "timeline" in analytics and not analytics["timeline"].empty:
        tl = analytics["timeline"]
        result["timeline"] = tl[tl["team"] == team]
    return result


def _write_sp_team_sheet(ws, metrics, detail, assignments):
    people, colors = _assign_colors(assignments)
    if not people:
        people = metrics.sort_values("total", ascending=False)["person"].tolist()

    for c in (1, 2):
        cell = ws.cell(row=1, column=c)
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.border = THIN_BOTTOM
    ws.cell(row=1, column=1, value="FN")
    ws.cell(row=1, column=2, value="Person")
    ws.freeze_panes = "A2"

    for i, (_, r) in enumerate(detail.iterrows()):
        row = i + 2
        fn = r["footnote"]
        ws.cell(row=row, column=1, value=fn)
        cell = ws.cell(row=row, column=2, value=r["modified_by"])
        f = _range_fill(fn, assignments, colors)
        if f:
            cell.fill = f

    col = 4
    ws.cell(row=3, column=col, value="Leaderboard").font = BOLD

    ws.cell(row=4, column=col, value="Name")
    for j, p in enumerate(people):
        cell = ws.cell(row=4, column=col + 1 + j, value=p)
        if p in colors:
            cell.fill = colors[p]

    mdict = {r["person"]: r for _, r in metrics.iterrows()}

    ws.cell(row=5, column=col, value="Total Final Contributions")
    for j, p in enumerate(people):
        m = mdict.get(p)
        ws.cell(row=5, column=col + 1 + j, value=int(m["total"]) if m is not None else 0)

    ws.cell(row=8, column=col, value="Breakdown").font = BOLD

    rows_data = [
        (9, "Contributions in Assigned", "in_assigned"),
        (10, "Contributions Outside Assigned", "outside_assigned"),
        (11, "Other Editor Edits to Assigned", "other_edits_to_assigned"),
    ]
    for rw, label, key in rows_data:
        ws.cell(row=rw, column=col, value=label)
        for j, p in enumerate(people):
            m = mdict.get(p)
            ws.cell(row=rw, column=col + 1 + j,
                    value=int(m[key]) if m is not None else 0)

    for c in range(col, col + 1 + len(people)):
        ws.cell(row=11, column=c).border = THIN_BOTTOM

    _finalize(ws)


def _write_ed_team_sheet(ws, metrics, detail, assignments, team_analytics=None):
    people, colors = _assign_colors(assignments)
    if not people:
        people = metrics.sort_values("total", ascending=False)["person"].tolist()

    for c in (1, 2):
        cell = ws.cell(row=1, column=c)
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.border = THIN_BOTTOM
    ws.cell(row=1, column=1, value="FN")
    ws.cell(row=1, column=2, value="Person")
    ws.freeze_panes = "A2"

    if not detail.empty:
        loc_order = {"below_line": 0, "comment": 1, "above_line": 2}
        d = detail.copy()
        d["_sort"] = d["location"].map(loc_order).fillna(1)
        d = d.sort_values(["footnote", "_sort", "person"]).reset_index(drop=True)

        for i, (_, r) in enumerate(d.iterrows()):
            row = i + 2
            fn = r["footnote"]
            loc = r["location"]
            if loc == "above_line":
                ws.cell(row=row, column=1, value="Above Line")
            else:
                ws.cell(row=row, column=1, value=fn if fn > 0 else 0)
            cell = ws.cell(row=row, column=2, value=r["person"])
            f = _range_fill(fn, assignments, colors)
            if f:
                cell.fill = f

    col = 4
    ws.cell(row=3, column=col, value="Leaderboard").font = BOLD

    ws.cell(row=4, column=col, value="Name")
    for j, p in enumerate(people):
        cell = ws.cell(row=4, column=col + 1 + j, value=p)
        if p in colors:
            cell.fill = colors[p]
    total_col = col + 1 + len(people)
    ws.cell(row=4, column=total_col, value="Total").font = BOLD

    mdict = {r["person"]: r for _, r in metrics.iterrows()}

    ws.cell(row=5, column=col, value="Total Contributions")
    t_total = 0
    for j, p in enumerate(people):
        m = mdict.get(p)
        val = int(m["total"]) if m is not None else 0
        ws.cell(row=5, column=col + 1 + j, value=val)
        t_total += val
    ws.cell(row=5, column=total_col, value=t_total)

    ws.cell(row=8, column=col, value="Breakdown").font = BOLD

    ws.cell(row=9, column=col, value="Contributions in Assigned")
    for j, p in enumerate(people):
        m = mdict.get(p)
        ws.cell(row=9, column=col + 1 + j, value=int(m["in_assigned"]) if m is not None else 0)

    ws.cell(row=10, column=col, value="Contributions Outside Assigned")
    for j, p in enumerate(people):
        m = mdict.get(p)
        ws.cell(row=10, column=col + 1 + j, value=int(m["outside_assigned"]) if m is not None else 0)

    ws.cell(row=11, column=col, value="Share of Total Edits")
    for j, p in enumerate(people):
        m = mdict.get(p)
        cell = ws.cell(row=11, column=col + 1 + j,
                       value=float(m["share_of_team"]) if m is not None else 0)
        cell.number_format = '0.00%'

    for c in range(col, total_col + 1):
        ws.cell(row=11, column=c).border = THIN_TOP

    # --- Analytics sections ---
    if team_analytics:
        cq = team_analytics.get("comment_quality")
        tl = team_analytics.get("timeline")
        cq_dict = {}
        tl_dict = {}
        if cq is not None and not cq.empty:
            cq_dict = {r["person"]: r for _, r in cq.iterrows()}
        if tl is not None and not tl.empty:
            tl_dict = {r["person"]: r for _, r in tl.iterrows()}

        if cq_dict:
            ws.cell(row=13, column=col, value="Comment Quality").font = BOLD

            ws.cell(row=14, column=col, value="Total Comments")
            for j, p in enumerate(people):
                c = cq_dict.get(p)
                ws.cell(row=14, column=col + 1 + j,
                        value=int(c["total_comments"]) if c is not None else 0)

            ws.cell(row=15, column=col, value="Rule Citations")
            for j, p in enumerate(people):
                c = cq_dict.get(p)
                ws.cell(row=15, column=col + 1 + j,
                        value=int(c["rule_citations"]) if c is not None else 0)

            ws.cell(row=16, column=col, value="Avg Comment Score")
            for j, p in enumerate(people):
                c = cq_dict.get(p)
                ws.cell(row=16, column=col + 1 + j,
                        value=round(float(c["avg_score"]), 2) if c is not None else 0)
            for c_idx in range(col, total_col + 1):
                ws.cell(row=16, column=c_idx).border = THIN_BOTTOM

        if tl_dict:
            act_row = 18 if cq_dict else 13
            ws.cell(row=act_row, column=col, value="Activity").font = BOLD

            ws.cell(row=act_row + 1, column=col, value="Active Days")
            for j, p in enumerate(people):
                t = tl_dict.get(p)
                ws.cell(row=act_row + 1, column=col + 1 + j,
                        value=int(t["active_days"]) if t is not None else 0)

            ws.cell(row=act_row + 2, column=col, value="Edits/Active Day")
            for j, p in enumerate(people):
                t = tl_dict.get(p)
                ws.cell(row=act_row + 2, column=col + 1 + j,
                        value=round(float(t["edits_per_active_day"]), 1) if t is not None else 0)
            for c_idx in range(col, total_col + 1):
                ws.cell(row=act_row + 2, column=c_idx).border = THIN_BOTTOM

    # --- Stacked bar chart ---
    _write_ed_stacked_bar(ws, metrics, people, col)

    _finalize(ws)


def _write_ed_stacked_bar(ws, metrics, people, col):
    chart_data_row = 23
    ws.cell(row=chart_data_row, column=col, value="")
    ws.cell(row=chart_data_row, column=col + 1, value="Below Line")
    ws.cell(row=chart_data_row, column=col + 2, value="Above Line")
    ws.cell(row=chart_data_row, column=col + 3, value="Comments")

    mdict = {r["person"]: r for _, r in metrics.iterrows()}
    for i, p in enumerate(people):
        r = chart_data_row + 1 + i
        m = mdict.get(p)
        ws.cell(row=r, column=col, value=p)
        ws.cell(row=r, column=col + 1, value=int(m["below_line"]) if m is not None else 0)
        ws.cell(row=r, column=col + 2, value=int(m["above_line"]) if m is not None else 0)
        ws.cell(row=r, column=col + 3, value=int(m["comments"]) if m is not None else 0)

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.title = "Contribution Breakdown"
    chart.y_axis.title = "Edits"
    chart.style = 10

    data = Reference(ws, min_col=col + 1, max_col=col + 3,
                     min_row=chart_data_row, max_row=chart_data_row + len(people))
    cats = Reference(ws, min_col=col,
                     min_row=chart_data_row + 1, max_row=chart_data_row + len(people))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 20
    chart.height = 12

    bar_colors = ["4472C4", "ED7D31", "A5A5A5"]
    for idx, color in enumerate(bar_colors):
        chart.series[idx].graphicalProperties.solidFill = color

    anchor = f"{get_column_letter(col)}{chart_data_row + len(people) + 2}"
    ws.add_chart(chart, anchor)


TIMELINE_COLORS = [
    "4472C4", "ED7D31", "A5A5A5", "FFC000", "5B9BD5",
    "70AD47", "264478", "9B57A0", "636363", "EB7E3A",
]


def _write_timeline_chart(wb, velocity_df):
    ws = wb.create_sheet(title="Edit Timeline")
    ws.sheet_properties.tabColor = TAB_COLORS["analytics"]

    totals = velocity_df.groupby("person")["edit_count"].sum()
    top_people = totals.nlargest(10).index.tolist()
    vel = velocity_df[velocity_df["person"].isin(top_people)].copy()

    pivot = vel.pivot_table(index="day", columns="person",
                            values="edit_count", aggfunc="sum", fill_value=0)
    pivot = pivot.sort_index()
    pivot = pivot[sorted(pivot.columns, key=lambda p: -totals[p])]

    ws.cell(row=1, column=1, value="Date")
    for j, person in enumerate(pivot.columns):
        ws.cell(row=1, column=j + 2, value=person)

    for i, (day, row_data) in enumerate(pivot.iterrows()):
        ws.cell(row=i + 2, column=1, value=str(day))
        for j, person in enumerate(pivot.columns):
            ws.cell(row=i + 2, column=j + 2, value=int(row_data[person]))

    _write_analytics_header(ws, ["Date"] + list(pivot.columns))

    num_rows = len(pivot)
    num_cols = len(pivot.columns)

    chart = LineChart()
    chart.title = "Edit Activity Over Time"
    chart.y_axis.title = "Edits"
    chart.x_axis.title = "Date"
    chart.style = 10
    chart.width = 28
    chart.height = 14

    data = Reference(ws, min_col=2, max_col=num_cols + 1,
                     min_row=1, max_row=num_rows + 1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=num_rows + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    for idx in range(min(num_cols, len(TIMELINE_COLORS))):
        chart.series[idx].graphicalProperties.line.solidFill = TIMELINE_COLORS[idx]
        chart.series[idx].graphicalProperties.line.width = 22000

    anchor_row = num_rows + 3
    ws.add_chart(chart, f"A{anchor_row}")
    _finalize(ws)


def _write_analytics_header(ws, headers):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.border = THIN_BOTTOM
    ws.freeze_panes = "A2"


def _write_analytics_sheets(wb, analytics):
    if not analytics:
        return

    if "timeline" in analytics and not analytics["timeline"].empty:
        ws = wb.create_sheet(title="Work Timeline")
        ws.sheet_properties.tabColor = TAB_COLORS["analytics"]
        tl = analytics["timeline"]
        headers = ["Person", "Team", "First Edit", "Last Edit", "Days Span",
                   "Active Days", "Total Edits", "Edits/Active Day"]
        _write_analytics_header(ws, headers)
        for i, (_, r) in enumerate(tl.iterrows()):
            ws.cell(row=i + 2, column=1, value=r["person"])
            ws.cell(row=i + 2, column=2, value=r["team"])
            ws.cell(row=i + 2, column=3, value=str(r["first_edit"])[:16])
            ws.cell(row=i + 2, column=4, value=str(r["last_edit"])[:16])
            ws.cell(row=i + 2, column=5, value=r["days_span"])
            ws.cell(row=i + 2, column=6, value=r["active_days"])
            ws.cell(row=i + 2, column=7, value=r["total_edits"])
            ws.cell(row=i + 2, column=8, value=r["edits_per_active_day"])
        _finalize(ws)

    if "velocity" in analytics and not analytics["velocity"].empty:
        ws = wb.create_sheet(title="Edit Velocity")
        ws.sheet_properties.tabColor = TAB_COLORS["analytics"]
        vel = analytics["velocity"]
        headers = ["Person", "Team", "Date", "Edit Count"]
        _write_analytics_header(ws, headers)
        for i, (_, r) in enumerate(vel.iterrows()):
            ws.cell(row=i + 2, column=1, value=r["person"])
            ws.cell(row=i + 2, column=2, value=r["team"])
            ws.cell(row=i + 2, column=3, value=str(r["day"]))
            ws.cell(row=i + 2, column=4, value=r["edit_count"])
        _finalize(ws)

        _write_timeline_chart(wb, vel)

    if "heatmap" in analytics and not analytics["heatmap"].empty:
        ws = wb.create_sheet(title="Footnote Heatmap")
        ws.sheet_properties.tabColor = TAB_COLORS["analytics"]
        hm = analytics["heatmap"]
        top_hm = pd.concat([
            group.nlargest(50, "total_edits")
            for _, group in hm.groupby("team")
        ]).reset_index(drop=True)
        headers = ["Team", "Footnote", "Total Edits", "Num Editors", "Edits/Editor"]
        _write_analytics_header(ws, headers)
        for i, (_, r) in enumerate(top_hm.iterrows()):
            ws.cell(row=i + 2, column=1, value=r["team"])
            ws.cell(row=i + 2, column=2, value=r["footnote"])
            ws.cell(row=i + 2, column=3, value=r["total_edits"])
            ws.cell(row=i + 2, column=4, value=r["num_editors"])
            ws.cell(row=i + 2, column=5, value=r["edits_per_editor"])
        _finalize(ws)

    if "overlap" in analytics and not analytics["overlap"].empty:
        ws = wb.create_sheet(title="Editor Overlap")
        ws.sheet_properties.tabColor = TAB_COLORS["analytics"]
        ov = analytics["overlap"]
        headers = ["Team", "Person A", "Person B", "Shared Footnotes", "Overlap Edits"]
        _write_analytics_header(ws, headers)
        for i, (_, r) in enumerate(ov.iterrows()):
            ws.cell(row=i + 2, column=1, value=r["team"])
            ws.cell(row=i + 2, column=2, value=r["person_a"])
            ws.cell(row=i + 2, column=3, value=r["person_b"])
            ws.cell(row=i + 2, column=4, value=r["shared_footnotes"])
            ws.cell(row=i + 2, column=5, value=r["overlap_edits"])
        _finalize(ws)

    if "comment_quality" in analytics and not analytics["comment_quality"].empty:
        ws = wb.create_sheet(title="Comment Quality")
        ws.sheet_properties.tabColor = TAB_COLORS["analytics"]
        cq = analytics["comment_quality"]
        headers = ["Person", "Team", "Total Comments", "Rule Citations",
                   "Substantive", "Low Effort", "Avg Score"]
        _write_analytics_header(ws, headers)
        for i, (_, r) in enumerate(cq.iterrows()):
            ws.cell(row=i + 2, column=1, value=r["person"])
            ws.cell(row=i + 2, column=2, value=r["team"])
            ws.cell(row=i + 2, column=3, value=r["total_comments"])
            ws.cell(row=i + 2, column=4, value=r["rule_citations"])
            ws.cell(row=i + 2, column=5, value=r["substantive"])
            ws.cell(row=i + 2, column=6, value=r["low_effort"])
            ws.cell(row=i + 2, column=7, value=r["avg_score"])
        _finalize(ws)

    if "comment_scores" in analytics and not analytics["comment_scores"].empty:
        ws = wb.create_sheet(title="Comment Detail")
        ws.sheet_properties.tabColor = TAB_COLORS["analytics"]
        cs = analytics["comment_scores"]
        headers = ["Person", "Team", "FN", "Category", "Score", "Text"]
        _write_analytics_header(ws, headers)
        for i, (_, r) in enumerate(cs.iterrows()):
            ws.cell(row=i + 2, column=1, value=r["person"])
            ws.cell(row=i + 2, column=2, value=r["team"])
            ws.cell(row=i + 2, column=3, value=r["footnote"])
            ws.cell(row=i + 2, column=4, value=r["category"])
            ws.cell(row=i + 2, column=5, value=r["score"])
            ws.cell(row=i + 2, column=6, value=r["text"])
        _finalize(ws)


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 45)
